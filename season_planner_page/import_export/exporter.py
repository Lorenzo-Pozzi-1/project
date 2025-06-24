"""
Excel exporter for scenario export.

Exports scenarios to Excel format with each scenario as a separate worksheet.
"""

from dataclasses import dataclass
from typing import List, Tuple, Any, Optional
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Font, PatternFill, Alignment
from PySide6.QtWidgets import QMessageBox, QFileDialog


@dataclass
class ExportConfig:
    """Configuration constants for Excel export."""
    MAX_SHEET_NAME_LENGTH = 31
    INVALID_SHEET_CHARS = ['\\', '/', '*', '[', ']', ':', '?']
    DEFAULT_SHEET_NAME = "Scenario"
    MAX_COLUMN_WIDTH = 50
    COLUMN_PADDING = 2
    
    COLUMNS = [
        "App #", "Date", "Product Type", "Product Name", "Rate",
        "Rate UOM", "Area", "Method", "AI Groups", "Field EIQ"
    ]
    
    # Styling constants
    HEADER_FONT = Font(bold=True)
    HEADER_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    METADATA_FONT = Font(bold=True)
    CENTER_ALIGNMENT = Alignment(horizontal='center')


class DataConverter:
    """Helper class for safe data type conversions."""
    
    @staticmethod
    def to_int(value: Any, default: Any = '') -> Any:
        """Safely convert value to int, return default if conversion fails."""
        if not value or not str(value).strip():
            return default
        try:
            return int(value)
        except (ValueError, TypeError):
            return value
    
    @staticmethod
    def to_float(value: Any, default: Any = '') -> Any:
        """Safely convert value to float, return default if conversion fails."""
        if not value or not str(value).strip():
            return default
        try:
            return float(value)
        except (ValueError, TypeError):
            return value
    
    @staticmethod
    def to_rounded_float(value: Any, decimals: int = 2, default: Any = '') -> Any:
        """Safely convert value to rounded float."""
        if not value or not str(value).strip():
            return default
        try:
            return round(float(value), decimals)
        except (ValueError, TypeError):
            return value
    
    @staticmethod
    def to_string(value: Any, default: str = '') -> str:
        """Safely convert value to string."""
        return str(value) if value is not None else default


class ExcelFormatter:
    """Handles Excel worksheet formatting."""
    
    def __init__(self, config: ExportConfig):
        self.config = config
    
    def format_worksheet(self, worksheet, header_row_num: int) -> None:
        """Apply formatting to the Excel worksheet."""
        try:
            self._format_headers(worksheet, header_row_num)
            self._auto_adjust_columns(worksheet)
            self._format_metadata(worksheet)
        except Exception as e:
            print(f"Warning: Could not apply Excel formatting: {e}")
    
    def _format_headers(self, worksheet, header_row_num: int) -> None:
        """Format header row with bold font and grey background."""
        for col in range(1, len(self.config.COLUMNS) + 1):
            cell = worksheet.cell(row=header_row_num, column=col)
            cell.font = self.config.HEADER_FONT
            cell.fill = self.config.HEADER_FILL
            cell.alignment = self.config.CENTER_ALIGNMENT
    
    def _auto_adjust_columns(self, worksheet) -> None:
        """Auto-adjust column widths based on content."""
        for column_cells in worksheet.columns:
            max_length = 0
            column_letter = column_cells[0].column_letter
            
            for cell in column_cells:
                try:
                    cell_value = str(cell.value) if cell.value is not None else ""
                    max_length = max(max_length, len(cell_value))
                except:
                    pass
            
            # Set column width with padding, capped at max width
            adjusted_width = min(max_length + self.config.COLUMN_PADDING, self.config.MAX_COLUMN_WIDTH)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def _format_metadata(self, worksheet) -> None:
        """Format metadata rows with bold labels."""
        for row in range(1, 3):  # First two metadata rows
            for col in range(1, len(self.config.COLUMNS) + 1):
                cell = worksheet.cell(row=row, column=col)
                if col in [1, 3, 5, 7, 9]:  # Label columns
                    cell.font = self.config.METADATA_FONT


class ScenarioDataWriter:
    """Handles writing scenario data to worksheet."""
    
    def __init__(self, config: ExportConfig, converter: DataConverter):
        self.config = config
        self.converter = converter
        self.header_row_num = None
    
    def write_scenario_to_worksheet(self, worksheet, scenario) -> None:
        """Write complete scenario data to worksheet."""
        current_row = 1
        current_row = self._write_metadata(worksheet, scenario, current_row)
        current_row = self._write_headers(worksheet, current_row)
        self._write_applications(worksheet, scenario.applications or [], current_row)
    
    def _write_metadata(self, worksheet, scenario, start_row: int) -> int:
        """Write scenario metadata to worksheet."""
        current_row = start_row
        
        # Calculate total Field EIQ for the scenario
        total_eiq = self._calculate_total_field_eiq(scenario)
        
        # Scenario name + total field EIQ row
        worksheet.cell(row=current_row, column=1, value="Scenario Name:")
        worksheet.cell(row=current_row, column=2, value=scenario.name or "Unnamed Scenario")
        worksheet.cell(row=current_row, column=3, value="Total Field Use EIQ Score:")
        worksheet.cell(row=current_row, column=4, value="=SUM(J5:J100)")

        current_row += 1
        
        # Metadata row with multiple fields
        metadata_items = self._get_metadata_items(scenario)
        for col_offset, (label, value) in enumerate(metadata_items):
            col_start = (col_offset * 2) + 1
            worksheet.cell(row=current_row, column=col_start, value=label)
            worksheet.cell(row=current_row, column=col_start + 1, value=value)
        
        current_row += 1
        return current_row + 1  # +1 for separator row
    
    def _calculate_total_field_eiq(self, scenario) -> float:
        """Calculate total Field EIQ for all applications in the scenario."""
        total_eiq = 0.0
        if hasattr(scenario, 'applications') and scenario.applications:
            for app in scenario.applications:
                field_eiq = getattr(app, 'field_eiq', None)
                if field_eiq is not None:
                    try:
                        total_eiq += float(field_eiq)
                    except (ValueError, TypeError):
                        pass  # Skip invalid values
        return round(total_eiq, 2)
    
    def _get_metadata_items(self, scenario) -> List[Tuple[str, Any]]:
        """Get metadata items for scenario."""
        return [
            ("Crop Year:", self.converter.to_int(getattr(scenario, 'crop_year', ''))),
            ("Grower:", self.converter.to_string(getattr(scenario, 'grower_name', ''))),
            ("Field:", self.converter.to_string(getattr(scenario, 'field_name', ''))),
            ("Field Area:", self._format_field_area(scenario)),
            ("Variety:", self.converter.to_string(getattr(scenario, 'variety', '')))
        ]
    
    def _format_field_area(self, scenario) -> str:
        """Format field area with unit."""
        field_area = getattr(scenario, 'field_area', 0) or 0
        field_area_uom = getattr(scenario, 'field_area_uom', 'acre') or 'acre'
        
        area_value = self.converter.to_float(field_area, field_area)
        return f"{area_value} {field_area_uom}"
    
    def _write_headers(self, worksheet, start_row: int) -> int:
        """Write column headers."""
        for col_idx, header in enumerate(self.config.COLUMNS, 1):
            worksheet.cell(row=start_row, column=col_idx, value=header)
        
        self.header_row_num = start_row
        return start_row + 1
    
    def _write_applications(self, worksheet, applications: List, start_row: int) -> None:
        """Write application data rows."""
        for i, app in enumerate(applications):
            row = start_row + i
            self._write_application_row(worksheet, row, i + 1, app)
    
    def _write_application_row(self, worksheet, row: int, app_num: int, app) -> None:
        """Write a single application row."""
        data = [
            app_num,  # App # - numeric
            getattr(app, 'application_date', ''),  # Date
            getattr(app, 'product_type', ''),  # Product Type
            getattr(app, 'product_name', ''),  # Product Name
            self.converter.to_float(getattr(app, 'rate', '')),  # Rate
            getattr(app, 'rate_uom', ''),  # Rate UOM
            self.converter.to_float(getattr(app, 'area', '')),  # Area
            getattr(app, 'application_method', ''),  # Method
            ', '.join(getattr(app, 'ai_groups', []) or []),  # AI Groups
            self.converter.to_rounded_float(getattr(app, 'field_eiq', ''))  # Field EIQ
        ]
        
        for col_idx, value in enumerate(data, 1):
            worksheet.cell(row=row, column=col_idx, value=value)


class FileHandler:
    """Handles file operations and user interactions."""
    
    @staticmethod
    def get_save_file_path(parent_widget) -> Optional[str]:
        """Get file path from user via file dialog."""
        file_path, _ = QFileDialog.getSaveFileName(
            parent_widget,
            "Save Excel Export",
            "*.xlsx",
            "Excel Files (*.xlsx);;All Files (*)"
        )
        
        if not file_path:
            return None
        
        # Ensure .xlsx extension
        if not file_path.lower().endswith('.xlsx'):
            file_path = file_path + '.xlsx'
        
        return file_path
    
    @staticmethod
    def show_success_message(parent_widget, file_path: str) -> None:
        """Show success message to user."""
        if parent_widget:
            QMessageBox.information(
                parent_widget,
                "Export Successful",
                f"Scenarios exported successfully!\n\nFile saved to:\n{file_path}"
            )
    
    @staticmethod
    def show_error_message(parent_widget, message: str) -> None:
        """Show error message to user."""
        if parent_widget:
            QMessageBox.critical(parent_widget, "Export Failed", message)


class ExcelScenarioExporter:
    """Main exporter class for scenario data to Excel format."""
    
    def __init__(self):
        """Initialize the exporter with configuration and helper objects."""
        self.config = ExportConfig()
        self.converter = DataConverter()
        self.formatter = ExcelFormatter(self.config)
        self.data_writer = ScenarioDataWriter(self.config, self.converter)
        self.file_handler = FileHandler()
    
    def export_scenarios(self, scenarios: List, parent_widget=None) -> Optional[str]:
        """
        Export scenarios to Excel file with user-selected location and filename.
        
        Args:
            scenarios: List of scenario objects to export
            parent_widget: Parent widget for message boxes and dialogs
            
        Returns:
            Path to created file or None if failed
        """
        try:
            file_path = self.file_handler.get_save_file_path(parent_widget)
            if not file_path:
                return None
            
            self._create_excel_file(scenarios, file_path)
            self.file_handler.show_success_message(parent_widget, file_path)
            return file_path
            
        except PermissionError:
            self._handle_error("Permission denied. The file may be open in another application.", parent_widget)
        except FileNotFoundError:
            self._handle_error("Could not access the specified directory.", parent_widget)
        except Exception as e:
            self._handle_error(f"Failed to export scenarios: {str(e)}", parent_widget)
        
        return None
    
    def _create_excel_file(self, scenarios: List, file_path: str) -> None:
        """Create Excel file with scenario data."""
        workbook = Workbook()
        
        # Remove default sheet
        if workbook.worksheets:
            workbook.remove(workbook.active)
        
        for scenario in scenarios:
            sheet_name = self._sanitize_sheet_name(scenario.name)
            worksheet = workbook.create_sheet(title=sheet_name)
            
            self.data_writer.write_scenario_to_worksheet(worksheet, scenario)
            self.formatter.format_worksheet(worksheet, self.data_writer.header_row_num)
        
        workbook.save(file_path)
    
    def _sanitize_sheet_name(self, name: str) -> str:
        """
        Sanitize scenario name for use as Excel worksheet name.
        
        Args:
            name: Original scenario name
            
        Returns:
            Sanitized name suitable for Excel worksheet
        """
        if not name:
            return self.config.DEFAULT_SHEET_NAME
        
        sanitized = name
        for char in self.config.INVALID_SHEET_CHARS:
            sanitized = sanitized.replace(char, '_')
        
        # Limit to Excel's maximum length
        if len(sanitized) > self.config.MAX_SHEET_NAME_LENGTH:
            sanitized = sanitized[:self.config.MAX_SHEET_NAME_LENGTH]
        
        return sanitized.strip() or self.config.DEFAULT_SHEET_NAME
    
    def _handle_error(self, message: str, parent_widget) -> None:
        """Handle and display error messages."""
        print(f"Export error: {message}")
        self.file_handler.show_error_message(parent_widget, message)