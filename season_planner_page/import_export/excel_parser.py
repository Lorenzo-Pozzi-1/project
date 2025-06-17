"""
Excel parser for scenario import with round-trip support.

Supports both:
1. Original external Excel format
2. Files exported by the application itself (round-trip compatibility)
"""

import os
from openpyxl import load_workbook
from datetime import datetime

from data.model_application import Application
from data.model_scenario import Scenario
from data.repository_product import ProductRepository


class ExcelScenarioParser:
    """Parser that supports both external and exported Excel formats."""
    
    def __init__(self):
        """Initialize the parser."""
        # UOM mapping from Excel format to application format
        self.uom_mapping = {
            "Fluid Ounce per Hundred Weight": "fl oz/cwt",
            "Ounce per Acre": "oz/acre",
            "Pounds per Hundred Weight": "lb/cwt",
            "Pint per Acre": "pt/acre",
            "Pounds per Acre": "lb/acre",
            "Fluid Ounce per Acre": "fl oz/acre",
            "Milliliters per 100 Kilograms": "ml/100kg",
            "Milliliter per Acre": "ml/acre",
            "Grams per Acre": "g/acre",
            "Kilograms per Acre": "kg/acre",
            "Gallons per Acre": "gal/acre",
            "Liquid Quart per Acre": "qt/acre",
            "Ounce per Hundred Weight": "oz/cwt",
            "Liter per Acre": "l/acre",
            "Milliliters per Hundred Weight": "ml/cwt",
            "Liter per Hectare": "l/acre",
            "Liters per 100 Meter Row": "l/100m",
            "Milliliter per Hectare": "ml/ha",
            "Kilograms per Hectares": "kg/ha",
            "Milliliters per 100 Meter Row": "ml/100m",
            "Imperial Gallons per Acre": "CA gal/acre",
            "Short Ton per Acre": "US ton/acre",
            "Grams per Ton": "g/metric ton",
            "Milliliters per Tonne": "ml/metric ton",
            "Milliliters per Kilograms": "ml/kg"
        }

        # Get product repository instance
        self.products_repo = ProductRepository.get_instance()
    
    def parse_file(self, file_path):
        """
        Parse an Excel file and return a Scenario object.
        Automatically detects format type.
        
        Args:
            file_path (str): Path to the Excel file
            
        Returns:
            tuple: (Scenario object, preview_info dict) or (None, None) if error
        """
        try:
            # Load workbook with openpyxl
            workbook = load_workbook(file_path, data_only=True)
            worksheet = workbook.active
            
            # Convert worksheet to list of lists for processing
            data_rows = self._worksheet_to_rows(worksheet)
            
            # Detect format type
            format_type = self._detect_format_type(data_rows)
            
            if format_type == "exported":
                return self._parse_exported_format(data_rows, file_path)
            else:
                return self._parse_external_format(data_rows, file_path)
                
        except Exception as e:
            print(f"Error parsing Excel file: {e}")
            return None, None
    
    def _worksheet_to_rows(self, worksheet):
        """Convert worksheet to list of lists, similar to pandas DataFrame structure."""
        data_rows = []
        for row in worksheet.iter_rows(values_only=True):
            # Convert None values to empty strings for consistency
            processed_row = [cell if cell is not None else '' for cell in row]
            data_rows.append(processed_row)
        
        # Remove completely empty rows from the end
        while data_rows and all(cell == '' or cell is None for cell in data_rows[-1]):
            data_rows.pop()
            
        return data_rows
    
    def _detect_format_type(self, data_rows):
        """
        Detect whether this is an exported file or external format.
        
        Args:
            data_rows: List of lists representing worksheet data
            
        Returns:
            str: "exported" or "external"
        """
        try:
            # Check for exported format signature
            # Look for "Scenario Name:" in first cell of first row
            if len(data_rows) > 0 and len(data_rows[0]) > 0:
                first_cell = str(data_rows[0][0]).strip()
                
                if first_cell.startswith("Scenario Name:"):
                    return "exported"
            
            # Additional check: look for exported format headers in any row
            exported_headers = ["App #", "Date", "Product Type", "Product Name", "Rate", "Rate UOM", "Area", "Method"]
            
            for row_idx in range(min(10, len(data_rows))):  # Check first 10 rows
                row_values = [str(val).strip() for val in data_rows[row_idx] if val != '']
                
                # Check if this row contains multiple exported headers
                header_matches = sum(1 for header in exported_headers if header in row_values)
                if header_matches >= 4:  # If we find at least 4 exported headers
                    return "exported"
            
            # Check for specific exported format pattern in row 2 (metadata row)
            if len(data_rows) > 1:
                row2_values = [str(val).strip() for val in data_rows[1] if val != '']
                # Look for metadata pattern: "Crop Year:", "Grower:", "Field:", etc.
                metadata_indicators = ["Crop Year:", "Grower:", "Field:", "Field Area:", "Variety:"]
                metadata_matches = sum(1 for indicator in metadata_indicators if any(indicator in val for val in row2_values))
                
                if metadata_matches >= 2:
                    return "exported"
            
            return "external"
            
        except Exception as e:
            print(f"Error detecting format: {e}")
            return "external"
    
    def _parse_exported_format(self, data_rows, file_path):
        """Parse files exported by the application itself."""
        try:
            # Extract scenario name from row 1
            scenario_name = self._extract_exported_scenario_name(data_rows)
            
            # Extract metadata from row 2
            metadata = self._extract_exported_metadata(data_rows)
            
            # Find the header row
            header_row_index = self._find_exported_header_row(data_rows)
            
            if header_row_index is None:
                raise ValueError("Could not find header row in exported format")
            
            # Extract applications data starting after header row
            applications_data = data_rows[header_row_index + 1:]
            header_row = data_rows[header_row_index]
            
            # Clean and process applications
            applications_dict_list = self._clean_exported_applications(applications_data, header_row)
            
            # Validate products
            product_validation = self._validate_products_exported(applications_dict_list)
            
            # Create scenario
            scenario = self._create_exported_scenario(applications_dict_list, scenario_name, metadata, product_validation)
            
            # Prepare preview info
            preview_info = {
                'total_rows': len(applications_dict_list),
                'headers': ["App #", "Date", "Product Type", "Product Name", "Rate", "Rate UOM", "Area", "Method", "AI Groups", "Field EIQ"],
                'sample_data': self._format_sample_applications_exported(applications_dict_list),
                'crop_year': scenario.crop_year if scenario else 'N/A',
                'grower_name': scenario.grower_name if scenario else 'N/A',
                'field_name': scenario.field_name if scenario else 'N/A',
                'field_area': scenario.field_area if scenario else 'N/A',
                'field_area_uom': scenario.field_area_uom if scenario else '',
                'variety': scenario.variety if scenario else 'N/A',
                'product_validation': product_validation
            }
            
            return scenario, preview_info
            
        except Exception as e:
            print(f"Error parsing exported format: {e}")
            return None, None
    
    def _parse_external_format(self, data_rows, file_path):
        """Parse external Excel files (original functionality)."""
        try:
            # Find where data ends (first completely empty row)
            data_end_idx = self._find_data_end(data_rows)
            
            # Extract just the application data
            applications_data = data_rows[:data_end_idx]
            
            # Convert to dict format assuming first row is headers
            if not applications_data:
                raise ValueError("No data found in Excel file")
            
            headers = [str(cell).strip() for cell in applications_data[0]]
            applications_dict_list = []
            
            for row_data in applications_data[1:]:
                # Skip completely empty rows
                if all(cell == '' or cell is None for cell in row_data):
                    continue
                    
                row_dict = {}
                for i, header in enumerate(headers):
                    if i < len(row_data):
                        value = row_data[i]
                        # Convert to appropriate type
                        if value == '' or value is None:
                            if header in ['Rate', 'Acres']:
                                row_dict[header] = 0.0
                            else:
                                row_dict[header] = ''
                        else:
                            row_dict[header] = value
                    else:
                        row_dict[header] = '' if header not in ['Rate', 'Acres'] else 0.0
                
                applications_dict_list.append(row_dict)
            
            # Clean the data
            applications_dict_list = self._clean_dataframe_equivalent(applications_dict_list)
            
            # Validate and match products
            product_validation = self._validate_products(applications_dict_list)
            
            # Create scenario from the data
            scenario = self._create_scenario(applications_dict_list, file_path, product_validation)
            
            # Prepare preview info
            preview_info = {
                'total_rows': len(applications_dict_list),
                'headers': headers,
                'sample_data': self._format_sample_applications(applications_dict_list),
                'crop_year': scenario.crop_year if scenario else 'N/A',
                'grower_name': scenario.grower_name if scenario else 'N/A',
                'field_name': scenario.field_name if scenario else 'N/A', 
                'field_area': scenario.field_area if scenario else 'N/A',
                'field_area_uom': scenario.field_area_uom if scenario else '',
                'variety': scenario.variety if scenario else 'N/A',
                'product_validation': product_validation
            }
            
            return scenario, preview_info
            
        except Exception as e:
            print(f"Error parsing external format: {e}")
            return None, None
    
    # Exported format parsing methods
    
    def _extract_exported_scenario_name(self, data_rows):
        """Extract scenario name from exported format row 1."""
        try:
            # Look for scenario name in first row
            if not data_rows or not data_rows[0]:
                return "Imported Scenario"
                
            first_row = data_rows[0]
            
            # Check if first cell contains "Scenario Name:"
            first_cell = str(first_row[0]).strip()
            if first_cell.startswith("Scenario Name:"):
                # Name should be in the second column
                if len(first_row) > 1:
                    name = str(first_row[1]).strip()
                    return name if name and name != "None" and name != "" else "Imported Scenario"
                else:
                    # Try to extract from the same cell after the colon
                    name_part = first_cell.replace("Scenario Name:", "").strip()
                    return name_part if name_part else "Imported Scenario"
            
            return "Imported Scenario"
            
        except Exception:
            return "Imported Scenario"
    
    def _extract_exported_metadata(self, data_rows):
        """Extract metadata from exported format row 2."""
        try:
            metadata = {
                'crop_year': '',
                'grower_name': '',
                'field_name': '',
                'field_area': '',
                'field_area_uom': '',
                'variety': ''
            }
            
            if len(data_rows) <= 1:
                return metadata
            
            # Row 2 contains metadata in label:value pairs
            row2 = data_rows[1]
            row2_values = [str(val) for val in row2]
            
            # Parse the metadata row looking for patterns like "Label:" followed by value
            for i, val in enumerate(row2_values):
                val_str = str(val).strip()
                
                if val_str.startswith("Crop Year:") and i + 1 < len(row2_values):
                    metadata['crop_year'] = self._parse_crop_year_exported(str(row2_values[i + 1]))
                elif val_str.startswith("Grower:") and i + 1 < len(row2_values):
                    metadata['grower_name'] = str(row2_values[i + 1]).strip()
                elif val_str.startswith("Field:") and i + 1 < len(row2_values):
                    metadata['field_name'] = str(row2_values[i + 1]).strip()
                elif val_str.startswith("Field Area:") and i + 1 < len(row2_values):
                    area_str = str(row2_values[i + 1]).strip()
                    metadata['field_area'], metadata['field_area_uom'] = self._parse_area_exported(area_str)
                elif val_str.startswith("Variety:") and i + 1 < len(row2_values):
                    metadata['variety'] = str(row2_values[i + 1]).strip()
            
            return metadata
            
        except Exception as e:
            print(f"Error extracting exported metadata: {e}")
            return {
                'crop_year': '',
                'grower_name': '',
                'field_name': '',
                'field_area': '',
                'field_area_uom': 'acre',
                'variety': ''
            }
    
    def _parse_crop_year_exported(self, year_str):
        """Parse crop year from exported format."""
        try:
            year_str = str(year_str).strip()
            if year_str and year_str != "None" and year_str != "":
                year = int(float(year_str))
                return year if year > 1900 else datetime.now().year
        except Exception:
            pass
        return datetime.now().year
    
    def _parse_area_exported(self, area_str):
        """Parse field area and UOM from exported format."""
        try:
            area_str = str(area_str).strip()
            if area_str and area_str != "None" and area_str != "":
                # Format is like "10.0 acre"
                parts = area_str.split()
                if len(parts) >= 2:
                    area = float(parts[0])
                    uom = " ".join(parts[1:])
                    return area, uom
                elif len(parts) == 1:
                    return float(parts[0]), "acre"
        except Exception:
            pass
        return 0.0, "acre"
    
    def _find_exported_header_row(self, data_rows):
        """Find the header row in exported format."""
        try:
            # Look for row containing "App #" and "Product Name"
            for i in range(len(data_rows)):
                row_values = [str(val).strip() for val in data_rows[i] if val != '']
                
                if "App #" in row_values and "Product Name" in row_values:
                    return i
            
            return None
            
        except Exception:
            return None
    
    def _clean_exported_applications(self, applications_data, header_row):
        """Clean applications data from exported format."""
        try:
            # Set proper column names from header row
            columns = []
            for col in header_row:
                if col != '' and col is not None and str(col).strip():
                    columns.append(str(col).strip())
                else:
                    columns.append(f"Unnamed_{len(columns)}")
            
            applications_dict_list = []
            
            for row_data in applications_data:
                # Skip completely empty rows
                if all(cell == '' or cell is None for cell in row_data):
                    continue
                
                # Ensure we don't have more columns than data
                num_cols = min(len(columns), len(row_data))
                
                row_dict = {}
                for i in range(num_cols):
                    col_name = columns[i]
                    value = row_data[i] if i < len(row_data) else ''
                    
                    # Fill empty values with appropriate defaults
                    if value == '' or value is None:
                        if col_name in ['Rate', 'Area', 'App #']:
                            row_dict[col_name] = 0.0
                        else:
                            row_dict[col_name] = ''
                    else:
                        row_dict[col_name] = value
                
                # Filter out rows where App # is not a valid number
                if "App #" in row_dict:
                    try:
                        app_num = float(row_dict["App #"])
                        if app_num <= 0:
                            continue
                    except (ValueError, TypeError):
                        continue
                
                applications_dict_list.append(row_dict)
            
            return applications_dict_list
            
        except Exception as e:
            print(f"Error cleaning exported applications: {e}")
            return []
    
    def _validate_products_exported(self, applications_dict_list):
        """Validate products from exported format."""
        return self._validate_products_by_column(applications_dict_list, "Product Name")
    
    def _create_exported_scenario(self, applications_dict_list, scenario_name, metadata, product_validation):
        """Create scenario from exported format data."""
        try:
            # Create new scenario
            scenario = Scenario(scenario_name)
            
            # Set metadata
            scenario.crop_year = metadata.get('crop_year', 2024)
            scenario.grower_name = metadata.get('grower_name', '')
            scenario.field_name = metadata.get('field_name', '')
            scenario.field_area = metadata.get('field_area', 0.0)
            scenario.field_area_uom = metadata.get('field_area_uom', 'acre')
            scenario.variety = metadata.get('variety', '')
            
            # Create applications from each row
            applications = []
            product_mapping = product_validation['product_mapping']
            
            for row_dict in applications_dict_list:
                product_name = str(row_dict.get('Product Name', '')).strip()
                
                # Skip if no product name
                if not product_name or product_name == 'None':
                    continue
                
                matched_product = product_mapping.get(product_name)
                
                if matched_product is not None:
                    # Use matched product data
                    app_data = {
                        'application_date': self._format_date_string(row_dict.get('Date', ''), is_exported_format=True),
                        'product_name': matched_product.product_name,
                        'product_type': matched_product.product_type,
                        'rate': self._safe_float(row_dict.get('Rate', 0)),
                        'rate_uom': str(row_dict.get('Rate UOM', '')).strip(),
                        'application_method': str(row_dict.get('Method', 'Broadcast')),
                        'area': self._safe_float(row_dict.get('Area', 0))
                    }
                else:
                    # Keep unmatched product data as-is
                    app_data = {
                        'application_date': self._format_date_string(row_dict.get('Date', ''), is_exported_format=True),
                        'product_name': product_name,
                        'product_type': str(row_dict.get('Product Type', '')).strip(),
                        'rate': self._safe_float(row_dict.get('Rate', 0)),
                        'rate_uom': str(row_dict.get('Rate UOM', '')).strip(),
                        'application_method': str(row_dict.get('Method', 'Broadcast')),
                        'area': self._safe_float(row_dict.get('Area', 0))
                    }
                
                applications.append(Application.from_dict(app_data))
            
            scenario.applications = applications
            return scenario
            
        except Exception as e:
            print(f"Error creating exported scenario: {e}")
            return None
    
    def _format_sample_applications_exported(self, applications_dict_list):
        """Format sample applications for preview from exported format."""
        if len(applications_dict_list) == 0:
            return "No applications found"
        
        sample_lines = []
        for i, row_dict in enumerate(applications_dict_list[:3]):
            date = row_dict.get('Date', 'N/A')
            product = row_dict.get('Product Name', 'N/A')
            rate = row_dict.get('Rate', 'N/A')
            uom = row_dict.get('Rate UOM', 'N/A')
            method = row_dict.get('Method', 'N/A')
            
            sample_lines.append(f"{i+1}. {date} - {product} @ {rate} {uom} ({method})")
        
        if len(applications_dict_list) > 3:
            sample_lines.append(f"... and {len(applications_dict_list) - 3} more applications")
        
        return "\n".join(sample_lines)
    
    # Original external format methods (adapted for openpyxl)
    
    def _validate_products(self, applications_dict_list):
        """Original product validation for external format."""
        return self._validate_products_by_column(applications_dict_list, 'Control Product')
    
    def _validate_products_by_column(self, applications_dict_list, product_column):
        """Validate products against the application's product repository."""
        validation_results = {
            'total_products': 0,
            'matched_products': 0,
            'unmatched_products': 0,
            'matched_list': [],
            'unmatched_list': [],
            'product_mapping': {}
        }
        
        # Get all available products from repository
        available_products = self.products_repo.get_filtered_products()
        available_product_names = {product.product_name.lower(): product for product in available_products}
        
        # Extract unique product names from data
        excel_products = set()
        for row_dict in applications_dict_list:
            product_name = row_dict.get(product_column, '')
            if product_name and str(product_name).strip() and str(product_name).strip() != "None":
                excel_products.add(str(product_name).strip())
        
        excel_products = list(excel_products)
        validation_results['total_products'] = len(excel_products)
        
        for excel_product in excel_products:
            excel_product_lower = excel_product.lower()
            
            if excel_product_lower in available_product_names:
                matched_product = available_product_names[excel_product_lower]
                validation_results['matched_products'] += 1
                validation_results['matched_list'].append(excel_product)
                validation_results['product_mapping'][excel_product] = matched_product
            else:
                validation_results['unmatched_products'] += 1
                validation_results['unmatched_list'].append(excel_product)
                validation_results['product_mapping'][excel_product] = None
        
        return validation_results
    
    def _find_data_end(self, data_rows):
        """Find the index where application data ends (first blank row)."""
        for idx, row in enumerate(data_rows):
            if all(cell == '' or cell is None for cell in row):
                return idx
        return len(data_rows)
    
    def _clean_dataframe_equivalent(self, applications_dict_list):
        """Clean the dict list by removing empty entries and standardizing data."""
        cleaned_list = []
        
        for row_dict in applications_dict_list:
            # Skip completely empty rows
            if all(value == '' or value is None for value in row_dict.values()):
                continue
            
            # Clean individual values
            cleaned_dict = {}
            for key, value in row_dict.items():
                if value == '' or value is None:
                    if key in ['Rate', 'Acres']:
                        cleaned_dict[key] = 0.0
                    else:
                        cleaned_dict[key] = ''
                else:
                    cleaned_dict[key] = value
            
            cleaned_list.append(cleaned_dict)
        
        return cleaned_list
    
    def _format_sample_applications(self, applications_dict_list):
        """Format sample applications for preview display (external format)."""
        if len(applications_dict_list) == 0:
            return "No applications found"
        
        sample_lines = []
        for i, row_dict in enumerate(applications_dict_list[:3]):
            date = row_dict.get('Application Date', 'N/A')
            product = row_dict.get('Control Product', 'N/A')
            rate = row_dict.get('Rate', 'N/A')
            uom = row_dict.get('Rate UOM', 'N/A')
            method = row_dict.get('Application Method', 'N/A')
            
            sample_lines.append(f"{i+1}. {date} - {product} @ {rate} {uom} ({method})")
        
        if len(applications_dict_list) > 3:
            sample_lines.append(f"... and {len(applications_dict_list) - 3} more applications")
        
        return "\n".join(sample_lines)

    def _create_scenario(self, applications_dict_list, file_path, product_validation):
        """Create a Scenario object from the parsed data (external format)."""
        base_name = os.path.splitext(os.path.basename(file_path))[0]
        scenario_name = f"{base_name}"
        
        scenario = Scenario(scenario_name)
        
        if len(applications_dict_list) > 0:
            first_row = applications_dict_list[0]
            
            scenario.crop_year = self._extract_crop_year(first_row.get('Crop Year', ''))
            scenario.grower_name = str(first_row.get('Grower Name', '')).strip()
            scenario.field_name = str(first_row.get('GrowerFieldName', ''))
            scenario.field_area = self._safe_float(first_row.get('Acres', 0))
            scenario.field_area_uom = 'acre'
            scenario.variety = str(first_row.get('Variety', ''))
        
        applications = []
        product_mapping = product_validation['product_mapping']
        
        for row_dict in applications_dict_list:
            excel_product_name = str(row_dict.get('Control Product', '')).strip()
            
            if not excel_product_name or excel_product_name == 'None':
                continue
            
            matched_product = product_mapping.get(excel_product_name)
            
            if matched_product is not None:
                app_data = {
                    'application_date': self._format_date_string(row_dict.get('Application Date', ''), is_exported_format=False),
                    'product_name': matched_product.product_name,
                    'product_type': matched_product.product_type,
                    'rate': self._safe_float(row_dict.get('Rate', 0)),
                    'rate_uom': self._map_uom(row_dict.get('Rate UOM', '')),
                    'application_method': str(row_dict.get('Application Method', 'Broadcast')),
                    'area': self._safe_float(row_dict.get('Acres', 0))
                }
            else:
                app_data = {
                    'application_date': self._format_date_string(row_dict.get('Application Date', ''), is_exported_format=False),
                    'product_name': excel_product_name,
                    'product_type': '',
                    'rate': self._safe_float(row_dict.get('Rate', 0)),
                    'rate_uom': self._map_uom(row_dict.get('Rate UOM', '')),
                    'application_method': str(row_dict.get('Application Method', 'Broadcast')),
                    'area': self._safe_float(row_dict.get('Acres', 0))
                }
            
            applications.append(Application.from_dict(app_data))
        
        scenario.applications = applications
        return scenario
    
    def _safe_float(self, value):
        """Safely convert value to float."""
        try:
            if value == '' or value is None:
                return 0.0
            return float(value)
        except (ValueError, TypeError):
            return 0.0
    
    def _map_uom(self, excel_uom):
        """Map Excel UOM format to application UOM format."""
        excel_uom_clean = str(excel_uom).strip()
        return self.uom_mapping.get(excel_uom_clean, excel_uom.strip())
    
    def _format_date_string(self, date_value, is_exported_format=False):
        """
        Format date value from Excel to a clean date string.
        
        Args:
            date_value: Date value from Excel
            is_exported_format: If True, preserve date as-is (already formatted)
        
        Returns:
            str: Formatted date string
        """
        if not date_value or str(date_value).strip() == '' or str(date_value) == 'None':
            return ''
        
        try:
            # For exported format (round-trip), preserve the date as-is since it was already formatted
            if is_exported_format:
                date_str = str(date_value).strip()
                # Just return the string as-is, it's already in the correct format
                return date_str
            
            # For external format, perform full date parsing and formatting
            if hasattr(date_value, 'strftime'):
                return date_value.strftime('%d/%m/%Y')
            
            date_str = str(date_value).strip()
            
            if ' ' in date_str and ':' in date_str:
                date_str = date_str.split(' ')[0]
            
            date_formats = [
                '%Y-%m-%d',
                '%m/%d/%Y',
                '%d/%m/%Y',
                '%m-%d-%Y',
                '%d-%m-%Y',
                '%Y/%m/%d',
            ]
            
            for fmt in date_formats:
                try:
                    parsed_date = datetime.strptime(date_str, fmt)
                    return parsed_date.strftime('%d/%m/%Y')
                except ValueError:
                    continue
            
            return date_str
            
        except Exception:
            return str(date_value).strip()
    
    def _extract_crop_year(self, crop_year_str):
        """Extract numeric year from crop year string like 'CY24'."""
        try:
            if isinstance(crop_year_str, str) and crop_year_str.startswith('CY'):
                year_suffix = int(crop_year_str[2:])
                return 2000 + year_suffix if year_suffix > 0 else 2024
            return 2024
        except:
            return 2024